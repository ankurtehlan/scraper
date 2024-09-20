import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

# Step 1: Fetch the webpage
url = "https://www.marutisuzuki.com/genuine-parts/alto-800/2012-till-present/lxi"  # Replace with the actual URL
response = requests.get(url)

# Check if the request was successful
if response.status_code == 200:
    print("Successfully fetched the page")
else:
    print("Failed to fetch the page")
    exit()

# Step 2: Parse the webpage using BeautifulSoup
soup = BeautifulSoup(response.content, 'html.parser')

# Adjust the class names based on the structure of the website
product_cards = soup.find_all('div', class_='sliderBox')

# Create lists to store the scraped data
part_numbers = []
part_names = []
mrps = []
image_urls = []

# Step 3: Extract data from each product card
for card in product_cards:
    # Extract part number
    part_number = card.find('p').strong.text 
    print(part_number)
    
    # Extract part name
    part_name = card.find('h3').text.strip()
    
    # Extract MRP
    mrp = card.find('div', class_='price').text.strip()
    
    # Extract image URL
    image_url = card.find('img')['src']
    
    # Append extracted data to the lists
    part_numbers.append(part_number)
    part_names.append(part_name)
    mrps.append(mrp)
    image_urls.append(image_url)
    
    print(f"Scraped: {part_number}, {part_name}, {mrp}, {image_url}")

# Step 4: Download the product images to a local directory
if not os.path.exists('images'):
    os.makedirs('images')

image_paths = []
for idx, img_url in enumerate(image_urls):
    img_data = requests.get(img_url).content
    img_filename = f"images/part_image_{idx}.jpg"
    
    # Save the image locally
    with open(img_filename, 'wb') as img_file:
        img_file.write(img_data)
    
    # Append the local image path to the list
    image_paths.append(img_filename)
    print(f"Downloaded image {idx + 1}")

# Step 5: Create a DataFrame with the scraped data
data = {
    'Part Number': part_numbers,
    'Part Name': part_names,
    'MRP': mrps,
    'Image': image_paths
}

df = pd.DataFrame(data)

# Step 6: Save the DataFrame to an Excel file (without images for now)
df.to_excel('scraped_parts.xlsx', index=False)

# Step 7: Open the Excel file using openpyxl to embed the images
wb = Workbook()
ws = wb.active

# Set the header row
headers = ['Part Number', 'Part Name', 'MRP', 'Image']
ws.append(headers)

# Set column widths and row heights for images
for col in ['A', 'B', 'C', 'D']:
    ws.column_dimensions[col].width = 30

# Write data and insert images into the cells
for i, row in df.iterrows():
    # Write the part number, part name, and MRP into cells
    ws.cell(row=i+2, column=1, value=row['Part Number'])
    ws.cell(row=i+2, column=2, value=row['Part Name'])
    ws.cell(row=i+2, column=3, value=row['MRP'])
    
    # Insert the image into the Excel file
    img = ExcelImage(row['Image'])
    img.height = 100  # Set the image height
    img.width = 100   # Set the image width
    
    # Adjust the row height for image display
    ws.row_dimensions[i+2].height = 100
    
    # Add the image to the 'Image' column (D)
    ws.add_image(img, f"D{i+2}")

# Step 8: Save the Excel file with images embedded
wb.save('scraped_parts_with_images.xlsx')
print("Excel file saved with images!")
